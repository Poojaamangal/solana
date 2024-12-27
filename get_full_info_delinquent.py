import openpyxl
from openpyxl import Workbook

def read_keys_from_file(file_path):
    with open(file_path, 'r') as file:
        keys = file.read().splitlines()
    return set(keys)  # Use a set for faster lookups

def process_keys(input_file, delinquent_file, output_file):
    # Read delinquent keys
    delinquent_keys = read_keys_from_file(delinquent_file)

    # Load the Excel file
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Create a new workbook for the common keys
    common_wb = Workbook()
    common_sheet = common_wb.active

    # Iterate through the rows and check if the key in column F is in the delinquent keys
    for row in sheet.iter_rows(values_only=True):
        key = row[2]  # Column F (0-indexed, so it's 5)
        if key in delinquent_keys:
            common_sheet.append(row)
 # Save the common keys workbook
    common_wb.save(output_file)
input_file = "C:/Users/pooja/WORK/delinquent_workbench/testnet_managed.xlsx"
delinquent_file = "C:/Users/pooja/WORK/delinquent_workbench/delinquent.txt"
output_file = "C:/Users/pooja/WORK/delinquent_workbench/common.xlsx"
process_keys(input_file, delinquent_file, output_file)

