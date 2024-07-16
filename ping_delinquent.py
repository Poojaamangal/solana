import openpyxl
import subprocess
from openpyxl import Workbook
from tqdm import tqdm

def ping_ip(ip):
    try:
        output = subprocess.check_output(["ping", "-c", "1", ip], universal_newlines=True)
        return True if "1 packets transmitted, 1 received" in output else False
    except Exception as e:
        return False

def process_ip_addresses(input_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    reachable_wb = Workbook()
    unreachable_wb = Workbook()

    reachable_sheet = reachable_wb.active
    unreachable_sheet = unreachable_wb.active

    rows = list(sheet.iter_rows(values_only=True))
    for row in tqdm(rows, desc="Processing IP addresses"):
        ip_address = row[0]  # Assuming the IP address is in the first column

        if ping_ip(ip_address):
            reachable_sheet.append(row)
        else:
            unreachable_sheet.append(row)

    reachable_wb.save("reachable.xlsx")
    unreachable_wb.save("unreachable.xlsx")

input_file = "common.xlsx"  # Replace with your input Excel file name
process_ip_addresses(input_file)
